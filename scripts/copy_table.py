# copy_table.py

import openpyxl
import csv

def save_to_csv(csv_filename, sheet):
    with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
        csv_writer = csv.writer(csvfile)

        # Extract the first row for headers
        try:
            headers = next(sheet.iter_rows(values_only=True))
            if headers is None:
                print("The sheet is empty. No headers found.")
                return
            csv_writer.writerow(headers)
        except StopIteration:
            print("The sheet is empty. No data to write.")
            return

        # Write the remaining rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            csv_writer.writerow(row)

    print(f"Data with headers successfully written to {csv_filename}")

def main():
    filename = 'data.xlsx'  # Ensure this path is correct relative to your script

    try:
        xlsx = openpyxl.load_workbook(filename, data_only=True)
    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")
        return
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    sheet_names = xlsx.sheetnames  # Updated method to get sheet names

    if 'ScenarioInfo' in sheet_names:
        tab = xlsx["ScenarioInfo"]
        csv_filename = 'TblScenario.csv'  # You can specify a different path if needed
        save_to_csv(csv_filename, tab)
    else:
        print("Sheet 'ScenarioInfo' not found in the workbook.")

if __name__ == "__main__":
    main()
