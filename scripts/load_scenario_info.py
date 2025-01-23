# python manage.py runscript load_scenario_info
import random
import openpyxl
from django.contrib.auth import get_user_model
from core.models import *
from django.contrib.auth.models import Group
from decimal import Decimal

def run():
    from email.utils import parseaddr
    filename = 'scripts/data.xlsx'

    xlsx = openpyxl.load_workbook(filename)
    sheet_names = xlsx.get_sheet_names()
    print (sheet_names)

    if 'ScenarioInfo' in sheet_names:
        scenario_infos = ScenarioInfo.objects.all()
        current_scenario_infos = []
        for scenario_info in scenario_infos:
            current_scenario_infos.append(scenario_info.name)

        tab = xlsx["ScenarioInfo"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value
            name = tab.cell(row = row_idx, column=2).value
            description = tab.cell(row = row_idx, column=3).value or ''
            data_revision = tab.cell(row = row_idx, column=4).value or 0
            condition = tab.cell(row = row_idx, column=5).value or 0
            type_id = tab.cell(row = row_idx, column=6).value or 0
            backout = tab.cell(row = row_idx, column=7).value or 0
            point_src = tab.cell(row = row_idx, column=9).value or 0
            atm_dep = tab.cell(row = row_idx, column=10).value or 0
            climate_change = tab.cell(row = row_idx, column=11).value or 0
            soil = tab.cell(row = row_idx, column=12).value or 0
            base_load = tab.cell(row = row_idx, column=13).value or 0


            if not name in current_scenario_infos and len(name) > 0:
                current_scenario_infos.append(name)
                new_scenario_info = ScenarioInfo(id=cast_id, name=name, description=description, data_revision=data_revision, condition=condition, type_id=type_id, backout=backout, point_src=point_src, atm_dep=atm_dep, climate_change=climate_change, soil=soil, base_load=base_load)
                new_scenario_info.save()

    if 'State' in sheet_names:
        states = State.objects.all()
        current_states = []
        for state in states:
            current_states.append(state.name)

        tab = xlsx["State"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            abbr = tab.cell(row = row_idx, column=2).value or ''
            name = tab.cell(row = row_idx, column=3).value or ''


            if not name in current_states and len(name) > 0:
                current_states.append(name)
                new_state = State(id=cast_id, abbreviation=abbr, name=name)
                new_state.save()
    if 'GeographicArea' in sheet_names:
        geographies = GeographicArea.objects.all()
        current_geographies = []
        for geography in geographies:
            current_geographies.append(f'{geography.name}_{geography.state}')

        tab = xlsx["GeographicArea"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            fips = tab.cell(row = row_idx, column=3).value or 0 
            name = tab.cell(row = row_idx, column=4).value or ''
            state = tab.cell(row = row_idx, column=5).value or ''
            ncounties = tab.cell(row = row_idx, column=6).value or 0


            if not f'{name}_{state}' in current_geographies and len(name) > 0:
                current_geographies.append(name)
                new_geography = GeographicArea(id=cast_id, fips=fips, name=name, state=state, ncounties=ncounties, county=0)
                new_geography.save()

    if 'County' in sheet_names:
        tab = xlsx["County"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            fips = int(tab.cell(row = row_idx, column=5).value)
            county = int(tab.cell(row = row_idx, column=1).value)
            geographic_area = GeographicArea.objects.get(fips=fips)
            geographic_area.county = county 
            geographic_area.save()

    if 'LandRiverSegment' in sheet_names:
        print("========================================================LandRiverSegment")
        land_river_segments = LandRiverSegment.objects.all()
        current_land_river_segments= []
        for lrs in land_river_segments:
            current_land_river_segments.append(f'{lrs.name}')

        tab = xlsx["LandRiverSegment"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            name = tab.cell(row = row_idx, column=2).value or ''
            fips = tab.cell(row = row_idx, column=6).value or 0 
            state_id = tab.cell(row = row_idx, column=7).value or 0
            county_id = tab.cell(row = row_idx, column=8).value or 0 
            acres = tab.cell(row = row_idx, column=12).value or 0.0
            state = State.objects.get(id=state_id)
            print('County: {}'.format(county_id))
            county = GeographicArea.objects.get(county=int(county_id))
            print("========================================================New Row")

            if not f'{name}' in current_land_river_segments and len(name) > 0:
                current_land_river_segments.append(name)
                print("========================================================Added")
                new_lrs= LandRiverSegment(id=cast_id, name=name, fips=fips, state=state, geographic_area=county, acres=acres)
                new_lrs.save()
    if 'BmpType' in sheet_names:
        bmp_types = BmpType.objects.all()
        current_bmp_types = []
        for bmp_type in bmp_types:
            current_bmp_types.append(bmp_type.name)

        tab = xlsx["BmpType"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            name = tab.cell(row = row_idx, column=2).value or ''


            if not name in current_bmp_types and len(name) > 0:
                current_bmp_types.append(name)
                new_bmp_type = BmpType(id=cast_id, name=name)
                new_bmp_type.save()

    if 'Sector' in sheet_names:
        sectors = Sector.objects.all()
        current_sectors = []
        for sector in sectors:
            current_sectors.append(sector.name)

        tab = xlsx["Sector"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            name = tab.cell(row = row_idx, column=2).value or ''


            if not name in current_sectors and len(name) > 0:
                current_sectors.append(name)
                new_sector = Sector(id=cast_id, name=name)
                new_sector.save()

    if 'BmpCategory' in sheet_names:
        bmp_categories = BmpCategory.objects.all()
        current_bmp_categories = []
        for bmp_category in bmp_categories:
            current_bmp_categories.append(bmp_category.name)

        tab = xlsx["BmpCategory"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            name = tab.cell(row = row_idx, column=2).value or ''


            if not name in current_bmp_categories and len(name) > 0:
                current_bmp_categories.append(name)
                new_bmp_category = BmpCategory(id=cast_id, name=name)
                new_bmp_category.save()

    if 'Agency' in sheet_names:
        agencies = Agency.objects.all()
        current_agencies = []
        for agency in agencies:
            current_agencies.append(agency.name)

        tab = xlsx["Agency"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            code = tab.cell(row = row_idx, column=2).value or ''
            name = tab.cell(row = row_idx, column=3).value or ''
            description = tab.cell(row = row_idx, column=4).value or ''


            if not name in current_agencies and len(name) > 0:
                current_agencies.append(name)
                new_agency = Agency(id=cast_id, code=code, name=name, description=description)
                new_agency.save()




    if 'AnimalGrp' in sheet_names:
        animals = AnimalGrp.objects.all()
        current_animals = []
        for animal in animals:
            current_animals.append(animal.name)

        tab = xlsx["AnimalGrp"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            name = tab.cell(row = row_idx, column=2).value or ''
            description = tab.cell(row = row_idx, column=3).value or ''


            if not name in current_animals and len(name) > 0:
                current_animals.append(name)
                new_animal = AnimalGrp(id=cast_id, name=name, description=description)
                new_animal.save()


    if 'LoadSrc' in sheet_names:
        load_srcs = LoadSrc.objects.all()
        current_load_srcs= []
        for load_src in load_srcs:
            current_load_srcs.append(load_src.name)

        tab = xlsx["LoadSrc"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            code = tab.cell(row = row_idx, column=3).value or ''
            name = tab.cell(row = row_idx, column=2).value or ''
            description = tab.cell(row = row_idx, column=4).value or ''


            if not name in current_load_srcs and len(name) > 0:
                current_load_srcs.append(name)
                new_load_src= LoadSrc(id=cast_id, code=code, name=name, description=description)
                new_load_src.save()

    if 'LoadSrc' in sheet_names:
        tab = xlsx["LoadSrc"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            load_src_id = tab.cell(row = row_idx, column=1).value or 0
            sector_id = tab.cell(row = row_idx, column=5).value or ''
            sector = Sector.objects.get(id=int(sector_id))
            load_src = LoadSrc.objects.get(id=int(load_src_id))
            load_src.sector = sector 
            load_src.save()



    if 'Bmp' in sheet_names:
        bmps = Bmp.objects.all()
        current_bmps = []
        for bmp in bmps:
            current_bmps.append(bmp.short_name)

        tab = xlsx["Bmp"]
        row_count = tab.max_row
        #0    BmpId	BmpShortName	BmpFullName	BmpCategoryId	BmpDetails	ParentBmpId	ParentBmpFullName	BmpGroupId	BmpTypeId
        #BmpId	BmpShortName	BmpFullName	BmpCategoryId	BmpDetails	ParentBmpId	ParentBmpFullName	BmpGroupId	BmpTypeId
        for row_idx in range(2, row_count + 1):
            cast_id = tab.cell(row = row_idx, column=1).value or 0
            short_name = tab.cell(row = row_idx, column=2).value or '' 
            name = tab.cell(row = row_idx, column=3).value or ''
            bmp_category = tab.cell(row = row_idx, column=4).value or 0
            bmp_type = tab.cell(row = row_idx, column=9).value or 0
            description = tab.cell(row = row_idx, column=5).value or ''


            if not short_name in current_bmps and len(short_name) > 0:
                current_bmps.append(short_name)
                new_bmp = Bmp(id=cast_id, short_name=short_name, name=name, bmp_category_id=bmp_category, bmp_type_id=bmp_type, description=description) 
                new_bmp.save()

    if 'BmpSector' in sheet_names:
        tab = xlsx["BmpSector"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            bmp_id = tab.cell(row = row_idx, column=1).value or 0
            sector_id = tab.cell(row = row_idx, column=2).value or ''
            print(bmp_id, sector_id)
            sector = Sector.objects.get(id=int(sector_id))
            bmp = Bmp.objects.get(id=int(bmp_id))
            print(bmp, sector)
            bmp.sector = sector
            bmp.save()


    current_cost_bmp_unit = {}
    if 'CostBmpUnit' in sheet_names:
        tab = xlsx["CostBmpUnit"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            bmp_id = tab.cell(row = row_idx, column=1).value or 0
            name = tab.cell(row = row_idx, column=2).value or ''
            current_cost_bmp_unit[bmp_id] = name

    if 'BmpCost' in sheet_names:
        cost_bmps = BmpCost.objects.all()
        current_cost_bmps = []
        for cost_bmp in cost_bmps:
            current_cost_bmps.append(cost_bmp.bmp_state)

        tab = xlsx["BmpCost"]
        row_count = tab.max_row
        for row_idx in range(2, row_count + 1):
            state_id = tab.cell(row=row_idx, column=1).value or 0
            bmp_id = tab.cell(row=row_idx, column=2).value or ''
            cost_per_unit_value = tab.cell(row=row_idx, column=3).value or ''
            bmp_state = f'{bmp_id}_{state_id}'
        
            if not bmp_state in current_cost_bmps and len(bmp_state) > 0:
                current_cost_bmps.append(bmp_state)
                
                # Fetch the Bmp instance
                try:
                    bmp_instance = Bmp.objects.get(id=bmp_id)
                except Bmp.DoesNotExist:
                    print(f"Bmp with id {bmp_id} does not exist.")
                    continue  # Skip to the next iteration if Bmp instance is not found
        
                try:
                    state_instance = State.objects.get(id=state_id)
                except Bmp.DoesNotExist:
                    print(f"Bmp with id {bmp_id} does not exist.")
                    continue  # Skip to the next iteration if Bmp instance is not found

                try:
                    cost_per_unit = Decimal(cost_per_unit_value) if cost_per_unit_value else Decimal('0.00')
                except (ValueError, InvalidOperation):
                    print(f"Invalid cost value at row {row_idx}: {cost_per_unit_value}")
                    continue  # Skip to the next iteration if cost value is invalid
                new_cost_bmp = BmpCost(
                    state=state_instance,
                    bmp=bmp_instance,  # Assign the Bmp instance
                    cost=cost_per_unit,
                    unit=current_cost_bmp_unit[bmp_id],
                    bmp_state=bmp_state
                )
        
                new_cost_bmp.save()
